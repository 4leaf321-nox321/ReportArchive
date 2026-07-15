"""엔티티 벌크 임포트 (데이터 채우기 v1) — 엑셀/CSV 시트로 객체+속성+관계 시딩.

관리자가 축(entity_type)을 고르고 시트를 올리면, 각 행을 그 축의 객체로 생성/갱신
하고(값=이름, 지정 열=속성), 관계열은 대상 축에서 값으로 객체를 찾아 링크한다.
전부 기존 서비스 재사용 — upsert_record_entity / create_entity(속성 검증) + add_relation
(관계·축 제약 검증). dry_run=True면 검증만(쓰기 없음)해 미리보기를 돌려준다.

행별로 관대하게 처리(한 행 실패가 전체를 막지 않음) — 쓰기 서비스가 행마다
커밋하므로 부분 성공이 그대로 남는다. 관계 대상을 못 찾으면 그 링크만 건너뛰고
경고를 남긴다(객체 자체는 생성).
"""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.modules.entities import services as ent
from app.modules.entities.models import EntityKindClass
from app.modules.entities.schemas import EntityCreate, EntityImportMapping

# 한 번에 처리할 행 상한(runaway 가드). 커넥터 페이지네이션/증분 동기화가 큰 배치를
# 넘길 수 있어 넉넉히 둔다(붙여넣기 UI 는 현실적으로 여기까지 안 감).
_MAX_ROWS = 20000


def parse_sheet(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    """업로드 파일(.csv/.xlsx) → (헤더, 행 dict 목록). 첫 행 = 헤더."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        raw = list(csv.reader(io.StringIO(text)))
    elif name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        raw = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise ValueError("지원 형식이 아닙니다: .csv 또는 .xlsx 파일을 올려주세요.")

    def _s(v) -> str:
        return "" if v is None else str(v).strip()

    # 완전히 빈 행 제거.
    raw = [r for r in raw if any(_s(c) for c in r)]
    if not raw:
        return [], []
    headers = [_s(h) for h in raw[0]]
    rows = []
    for r in raw[1:]:
        rows.append({h: (_s(r[i]) if i < len(r) else "") for i, h in enumerate(headers)})
    return headers, rows


def run_import(
    db: Session,
    *,
    mapping: EntityImportMapping,
    rows: list[dict],
    creator_user_id: int,
    dry_run: bool,
) -> dict:
    """행 목록 → {summary, rows}. dry_run 이면 검증만(쓰기 없음)."""
    type_row = ent.get_type(db, mapping.type_id)
    if type_row is None:
        raise ValueError(f"축을 찾을 수 없습니다: {mapping.type_id}")
    if len(rows) > _MAX_ROWS:
        raise ValueError(f"한 번에 최대 {_MAX_ROWS}행까지 가능합니다(현재 {len(rows)}행).")

    # 관계열의 대상 축을 미리 resolve(열마다 1회).
    rel_cols = [(rc, ent.get_type_by_slug(db, rc.target_type))
                for rc in mapping.relation_columns]
    is_record = type_row.kind_class == EntityKindClass.record

    # entity_ref 속성 — 표에 이름/별칭으로 입력한 값을 대상 축에서 찾아 id 로
    # 치환한다(관계열과 동일한 이름 해소; 속성 검증은 id 만 받으므로). 대상 축이
    # 지정된(ref_type_slug) 속성만 해소한다 — 미지정 자유참조는 id 입력이 필요.
    # 축 조회는 여기서 1회(키 → (라벨, 대상 EntityType|None)).
    ref_prop_axes = {}
    for d in ent.list_property_defs(
        db, owner_kind="entity_type", owner_id=type_row.id
    ):
        if d.data_type == "entity_ref" and d.ref_type_slug:
            ref_prop_axes[d.key] = (
                d.label or d.key,
                ent.get_type_by_slug(db, d.ref_type_slug),
                d.ref_type_slug,
            )

    out, counts = [], {"create": 0, "update": 0, "error": 0,
                       "linked": 0, "link_unresolved": 0}

    for idx, row in enumerate(rows, start=1):
        value = (row.get(mapping.value_column) or "").strip()
        if not value:
            out.append(_row(idx, "", "error", ["값(이름) 칸이 비었습니다."]))
            counts["error"] += 1
            continue

        # 코드 매칭(안정 식별자) — code_column 이 있으면 이름 대신 코드로 객체를 식별해
        # 이름이 흔들려도 재동기화 시 중복을 만들지 않는다(L1', 커넥터 code 매칭).
        code = None
        if mapping.code_column:
            code = (row.get(mapping.code_column) or "").strip() or None

        props = {}
        ref_errors = []
        for header, key in mapping.property_columns.items():
            cell = (row.get(header) or "").strip()
            if not cell:
                continue
            if key in ref_prop_axes:
                # entity_ref 속성 — 이름/별칭을 대상 축에서 찾아 id 로 치환.
                label, ref_axis, ref_slug = ref_prop_axes[key]
                if ref_axis is None:
                    ref_errors.append(f"{label}: 대상 축 '{ref_slug}' 을 찾을 수 없습니다")
                    continue
                hit = ent.resolve_existing(db, type_id=ref_axis.id, value=cell)
                if hit is None:
                    ref_errors.append(
                        f"{label}: '{cell}' 에 해당하는 '{ref_axis.label}' 객체가 없습니다"
                    )
                    continue
                props[key] = hit.id
            else:
                props[key] = cell

        # entity_ref 이름 해소 실패는 그 행을 error 로(무엇을 못 찾았는지 명시).
        if ref_errors:
            out.append(_row(idx, value, "error", ["속성 오류: " + "; ".join(ref_errors)]))
            counts["error"] += 1
            continue

        existing = ent.resolve_existing(
            db, type_id=type_row.id, value=value, code=code)
        status = "update" if existing else "create"

        # 속성 검증(쓰기 전) — 실패하면 그 행은 error.
        try:
            ent.validate_properties(db, type_row, props)
        except ValueError as exc:
            out.append(_row(idx, value, "error", [f"속성 오류: {exc}"]))
            counts["error"] += 1
            continue

        # 관계열 미리보기 — 대상 resolve 여부.
        rel_view = []
        for rc, target_type in rel_cols:
            tval = (row.get(rc.column) or "").strip()
            if not tval:
                continue
            if target_type is None:
                rel_view.append({"column": rc.column, "relation": rc.relation,
                                 "target": tval, "resolved": False,
                                 "message": f"대상 축 없음: {rc.target_type}"})
                continue
            # 관계 대상도 코드로 매칭 가능(match_key='code' 면 값을 코드로 간주).
            rc_code = tval if getattr(rc, "match_key", "value") == "code" else None
            found = ent.resolve_existing(db, type_id=target_type.id, value=tval, code=rc_code)
            rel_view.append({"column": rc.column, "relation": rc.relation,
                             "target": tval, "resolved": found is not None,
                             "message": None if found else "대상 객체 없음 — 링크 건너뜀"})

        messages = []
        if not dry_run:
            try:
                if is_record:
                    entity = ent.upsert_record_entity(
                        db, axis_slug=type_row.slug, name=value,
                        properties=props, code=code,
                        creator_user_id=creator_user_id)
                elif existing is not None:
                    # reference 축 기존 — 재사용(+빈 code 보강). 속성은 기존 정책상 유지.
                    if code and not existing.code:
                        existing.code = code
                        db.commit()
                    entity = existing
                else:
                    entity = ent.create_entity(
                        db, EntityCreate(type_id=type_row.id, value=value,
                                         code=code, properties=props or None),
                        creator_user_id=creator_user_id)
            except ValueError as exc:
                out.append(_row(idx, value, "error", [f"저장 실패: {exc}"], rel_view))
                counts["error"] += 1
                continue
            # 관계 링크 — 대상 찾은 것만. imported=src, target=dst.
            for rc, target_type in rel_cols:
                tval = (row.get(rc.column) or "").strip()
                if not tval or target_type is None:
                    continue
                rc_code = tval if getattr(rc, "match_key", "value") == "code" else None
                target = ent.resolve_existing(db, type_id=target_type.id, value=tval, code=rc_code)
                if target is None:
                    continue
                try:
                    ent.add_relation(db, src=entity, dst=target,
                                     relation=rc.relation,
                                     creator_user_id=creator_user_id)
                    counts["linked"] += 1
                except ValueError as exc:
                    messages.append(f"관계 '{rc.relation}'→{tval}: {exc}")

        counts[status] += 1
        counts["link_unresolved"] += sum(1 for r in rel_view if not r["resolved"])
        eid = entity.id if (not dry_run and entity is not None) else None
        out.append(_row(idx, value, status, messages, rel_view, entity_id=eid))

    return {
        "summary": {**counts, "total": len(rows), "committed": not dry_run},
        "rows": out,
    }


def _row(idx, value, status, messages, relations=None, entity_id=None) -> dict:
    return {"row": idx, "value": value, "status": status,
            "messages": messages, "relations": relations or [],
            "entity_id": entity_id}
